"""
This module reads an Excel file and allows the user to get all the valid worksheet names,
get the 1st line in a worksheet and iterate over the rest of the worksheet row by row
(next_row). The returned row is a hash which contains only the keys that are defined in
a configuration file.

This module depends on openpyxl and pyyaml.
"""

from __future__ import print_function
import sys
from openpyxl import load_workbook
import yaml
from Reader import Reader

WORKSHEETS_KEY_NAME = 'worksheets'
REQUIRED_HEADERS_KEY_NAME = 'required'
OPTIONAL_HEADERS_KEY_NAME = 'optional'

class XLSReader(Reader):
    """
    Reader for Excel file for the fields from worksheets defined in a configuration file
    """

    def __init__(self, xls_filename, conf_filename):
        """
        Constructor

        :param xls_filename: Excel file path
        :type xls_filename: basestring
        :param conf_filename: configuration file path
        :type conf_filename: basestring
        """
        with open(conf_filename, 'r') as conf_file:
            self.xls_conf = yaml.load(conf_file)
        self.workbook = load_workbook(xls_filename, read_only=True)
        self.worksheets = None
        self._active_worksheet = None
        self.row_offset = {}
        self.headers = {}
        self.valid = None

    def __iter__(self):
        return self

    @property
    def active_worksheet(self):
        return self._active_worksheet

    @active_worksheet.setter
    def active_worksheet(self, worksheet):
        self._active_worksheet = worksheet

    def valid_worksheets(self):
        """
        Get the list of the names of worksheets which have all the configured required headers

        :return: list of valid worksheet names in the Excel file
        :rtype: list
        """
        if self.worksheets is not None:
            return self.worksheets

        self.worksheets = []
        sheet_titles = self.workbook.sheetnames
        for title in self.xls_conf[WORKSHEETS_KEY_NAME]:
            # Check worksheet exists
            if title not in sheet_titles:
                continue

            # Check number of rows
            worksheet = self.workbook[title]
            if worksheet.max_row < 2:
                continue

            # Check required headers are present
            self.headers[title] = [cell.value if cell.value is None else cell.value.strip()
                                   for cell in worksheet[1]]
            required_headers = self.xls_conf[title].get(REQUIRED_HEADERS_KEY_NAME, [])
            if set(required_headers) <= set(self.headers[title]): # issubset
                self.worksheets.append(title)
            else:
                print('Worksheet '+title+' does not have all the required headers!', file=sys.stderr)
                self.valid = False

        return self.worksheets

    def get_valid_conf_keys(self):
        """
        :return: the list of valid worksheet names
        :rtype: list
        """
        return self.valid_worksheets()

    def set_current_conf_key(self, current_key):
        """
        Set the active_worksheet with value in $current_key

        :param current_key: the name of the worksheet
        :type current_key:  basestring
        :return: nothing
        :rtype: void
        """
        self.active_worksheet = current_key

    def is_valid(self):
        """
        Check that is all the worksheets contain required headers

        :return: True if all the worksheets contain required headers. False otherwise
        :rtype: bool
        """
        if self.valid is None:
            self.valid = True
            self.valid_worksheets()

        return self.valid


    def get_current_headers(self):
        """
        Retrieve the list of worksheets that have all the required headers

        :return: the list of valid worksheet names in the Excel file
        :rtype: list
        """
        worksheets = self.valid_worksheets()
        current_worksheet = self.active_worksheet
        if current_worksheet not in worksheets:
            print('Worksheet '+current_worksheet+' is not available or not valid!',
                  file=sys.stderr)
            return []

        return [x for x in self.headers[current_worksheet] if x is not None]

    def next(self):
        """
        Retrieve next data row

        :param worksheet: the name of the worksheet
        :type worksheet: basestring
        :return: A hash containing all the REQUIRED and OPTIONAL fields as keys
                and the corresponding data as values
        :rtype: dict
        """
        if self.worksheets is None:
            self.valid_worksheets()

        worksheet = self.active_worksheet
        if worksheet is None:
            print('No worksheet is specified!', file=sys.stderr)
            raise StopIteration

        if worksheet not in self.worksheets:
            print('Worksheet ' + worksheet + ' is not valid!', file=sys.stderr)
            raise StopIteration

        if worksheet not in self.row_offset:
            self.row_offset[worksheet] = 1
        self.row_offset[worksheet] += 1

        required_headers = self.xls_conf[worksheet].get(REQUIRED_HEADERS_KEY_NAME, [])
        optional_headers = self.xls_conf[worksheet].get(OPTIONAL_HEADERS_KEY_NAME, [])

        for row in self.workbook[worksheet].iter_rows(min_row=self.row_offset[worksheet]):
            num_cells = 0
            for cell in row:
                num_cells += 1

            data = {}
            has_notnull = False
            for header in required_headers+optional_headers:
                header_index = num_cells
                if header in self.headers[worksheet]:
                    header_index = self.headers[worksheet].index(header)
                if header_index >= num_cells:
                    data[header] = None
                    continue

                cell = row[header_index]
                if cell.value is not None:
                    has_notnull = True

                data[header] = cell.value

            if has_notnull:
                data['row_num'] = self.row_offset[worksheet]
                return data

            # no data on this row, continue to next
            self.row_offset[worksheet] += 1

        raise StopIteration
