from openpyxl import load_workbook
import os
import argparse

import check_samples
import sys
sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))) + os.path.sep + "xls2xml")
from xls2xml import xls2xml


def is_cell_value_empty(cell_value):
    return cell_value == "None" or cell_value == ""


def get_sample_names(eva_sample_sheet):
    sample_names = []
    num_rows = eva_sample_sheet.max_row
    num_cols = eva_sample_sheet.max_column

    for i in range(1, num_rows + 1):
        for j in range(1, num_cols + 1):
            if str(eva_sample_sheet.cell(i, j).value).strip().lower() == "sample name":
                first_sample_accession = str(eva_sample_sheet.cell(i+1, j-3).value).strip()
                first_sample_name = str(eva_sample_sheet.cell(i + 1, j).value).strip()
                if not is_cell_value_empty(first_sample_accession) and not is_cell_value_empty(first_sample_name):
                    raise Exception("ERROR: Both Novel Sample Names and Sample Accessions are present "
                                    "in the Metadata sheet. Only one of these should be present!")
                if is_cell_value_empty(first_sample_accession) and is_cell_value_empty(first_sample_name):
                    raise Exception("ERROR: Either Novel Sample Names or Sample Accessions should be present "
                                    "in the Metadata sheet!")
                # Use Biosample accessions by default
                if not is_cell_value_empty(first_sample_accession):
                    j -= 3
                while i <= num_rows:
                    i += 1
                    sample_name_value = str(eva_sample_sheet.cell(i, j).value).strip()
                    if is_cell_value_empty(sample_name_value):
                        continue
                    sample_names.append(sample_name_value)
                return sample_names

    raise Exception("Could not find sample names in the sheet: Sample")


def rewrite_samples_tab(eva_metadata_sheet):
    eva_metadata_sheet_copy = os.path.dirname(os.path.realpath(__file__)) + os.path.sep + \
                              ".".join(os.path.basename(eva_metadata_sheet).split(".")[:-1]) + \
                              "_with_sample_names_tab.xlsx"
    metadata_wb = load_workbook(eva_metadata_sheet, data_only=True)

    if "Sample" not in metadata_wb.sheetnames:
        raise Exception("Sample tab could not be found in the EVA metadata sheet: " + eva_metadata_sheet)
    else:
        sample_names = get_sample_names(metadata_wb['Sample'])
        metadata_wb.save(eva_metadata_sheet_copy)
        metadata_wb.close()
        metadata_wb_copy = load_workbook(eva_metadata_sheet_copy)
        if "Sample_Names" not in metadata_wb_copy.sheetnames:
            sample_name_sheet = metadata_wb_copy.create_sheet("Sample_Names")
        else:
            sample_name_sheet = metadata_wb_copy["Sample_Names"]
        sample_name_sheet.cell(1, 1).value = "Sample Name"
        row_index = 2
        for sample_name in sample_names:
            sample_name_sheet.cell(row_index, 1).value = sample_name
            row_index += 1
        metadata_wb_copy.save(eva_metadata_sheet_copy)
        metadata_wb_copy.close()

    return eva_metadata_sheet_copy


arg_parser = argparse.ArgumentParser(
        description='Transform and output validated data from an excel file to a XML file')
arg_parser.add_argument('--metadata-file', required=True, dest='metadata_file',
                        help='EVA Submission Metadata Excel sheet')
arg_parser.add_argument('--file-path', required=True, dest='filepath',
                        help='Path to the directory in which submitted files can be found')

args = arg_parser.parse_args()
file_path = args.filepath
metadata_file = args.metadata_file

data_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
xls_conf = data_dir + os.path.sep + "tests/data/EVA_xls2xml_v2.conf"
xls_schema = data_dir + os.path.sep + "tests/data/EVA_xls2xml_v2.schema"
xslt_filename = data_dir + os.path.sep + "tests/data/EVA_xls2xml_v2.xslt"

file_xml = os.path.splitext(metadata_file)[0] + ".file.xml"
sample_xml = os.path.splitext(metadata_file)[0] + ".sample.xml"

rewritten_metadata_file = rewrite_samples_tab(metadata_file)
xls2xml.convert_xls_to_xml(xls_conf, ["Files"], xls_schema, xslt_filename, rewritten_metadata_file, file_xml)
xls2xml.convert_xls_to_xml(xls_conf, ["Sample_Names"], xls_schema, xslt_filename, rewritten_metadata_file, sample_xml)
check_samples.get_sample_diff(file_path, file_xml, sample_xml, submission_type="eva")
